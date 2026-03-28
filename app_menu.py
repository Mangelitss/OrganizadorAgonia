import sys
import os
import json
import datetime
import webbrowser
import shutil  
import requests # AÑADIDO PARA LA CONEXIÓN A FIREBASE
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry

# Importamos las lógicas independientes
from logica_trono import cargar_datos
from logica_miercoles import generar_cuadrillas_miercoles, generar_html_miercoles
from logica_viernes import generar_cuadrillas_viernes, generar_html_viernes
from logica_ensayos import generar_html_ensayo
from logica_informes import crear_html_informe
from logica_licencia import comprobar_licencia
from logica_calendario import cargar_eventos, guardar_eventos, generar_html_calendario
from logica_viacrucis import generar_datos_viacrucis, generar_html_viacrucis
from logica_personalizada import generar_datos_personalizados, generar_html_personalizado

# ==========================================
# CONFIGURACIÓN Y COLORES
# ==========================================
VERSION_ACTUAL = "1.5" # <--- VERSIÓN ACTUAL DEL PROGRAMA

CONFIG = {
    "peso_trono_kg": 2000,
    "peso_cruz_kg": 200,      
    "limite_peso_persona": 90,
    "plazas_turno_a": 36,
    "archivo_datos": "datos.json"
}

# --- URL DE LA BASE DE DATOS EN LA NUBE (FIREBASE) ---
FIREBASE_URL = "https://licencias-gestor-cofrade-default-rtdb.europe-west1.firebasedatabase.app/censo_oficial.json"

# Paleta Mayordomía "Modern UI"
C_MORADO = "#4F1243"
C_MORADO_HOVER = "#7a1b67"
C_MORADO_BG = "#2a0a23"
C_ORO = "#D1B514"
C_ORO_HOVER = "#97820B"
C_BLANCO = "#ffffff"
C_GRIS_FONDO = "#f0f2f5"
C_TEXTO = "#333333"
C_ROJO = "#360928"

# ==========================================
# CLASE PRINCIPAL DE LA INTERFAZ
# ==========================================
class GestorCofradeAPP:
    def __init__(self, root):
        self.root = root        

        # 1. OCULTAMOS LA VENTANA PRINCIPAL AL ARRANCAR
        self.root.withdraw()
        
        try:
            self.root.iconbitmap("icono.ico")
        except:
            pass

        self.root.title("Gestor de Turnos Cristo de la Agonía V1.0 - Ntro. Padre Jesús Nazareno")
        self.root.geometry("1366x720")
        self.root.configure(bg=C_GRIS_FONDO)
        # HACEMOS LA VENTANA PRINCIPAL REESCALABLE
        self.root.resizable(True, True)

        # Variables de estado
        self.current_frame = None
        self.frames = {}
        
        # Contenedores Principales
        self.frame_menu = tk.Frame(self.root, bg=C_MORADO, width=250)
        self.frame_menu.pack(side=tk.LEFT, fill=tk.Y)
        self.frame_menu.pack_propagate(False) 
        
        self.frame_main = tk.Frame(self.root, bg=C_GRIS_FONDO)
        self.frame_main.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.crear_menu_lateral()
        self.crear_pantallas()
        self.mostrar_pantalla("Inicio")

        # F11 — alternar pantalla completa
        self._pantalla_completa = False
        def toggle_fullscreen(event=None):
            self._pantalla_completa = not self._pantalla_completa
            self.root.attributes('-fullscreen', self._pantalla_completa)
        self.root.bind("<F11>", toggle_fullscreen)

        # 2. VERIFICAMOS LA SESIÓN:
        self.verificar_sesion()

    def verificar_sesion(self):
        import os, json, base64
        archivo_sesion = "licencia_local.json"
        
        if os.path.exists(archivo_sesion):
            try:
                with open(archivo_sesion, "r") as f:
                    datos = json.load(f)
                    user = datos.get("usuario")
                    # Descodificamos la contraseña ofuscada
                    pwd = base64.b64decode(datos.get("password", "").encode()).decode()
                
                valido, mensaje = comprobar_licencia(user, pwd)
                
                if valido:
                    self.root.deiconify()
                    return
                elif "conexión" in mensaje.lower() or "internet" in mensaje.lower():
                    # Sin internet pero con sesión guardada: permitir acceso en modo offline
                    self.root.deiconify()
                    messagebox.showwarning(
                        "Modo sin conexión",
                        "⚠️ No se ha podido verificar la licencia online.\n\n"
                        "Se ha concedido acceso con la sesión guardada localmente.\n"
                        "Algunas funciones online pueden no estar disponibles."
                    )
                    return
            except Exception:
                pass 
                
        self.pedir_login()

    def pedir_login(self):
        self.ventana_login = tk.Toplevel(self.root)
        self.ventana_login.title("Activación de Licencia - Gestor Cofrade")
        self.ventana_login.geometry("450x600") 
        self.ventana_login.configure(bg=C_MORADO) 
        self.ventana_login.resizable(True, True)
        self.ventana_login.protocol("WM_DELETE_WINDOW", self.root.destroy)
        
        try:
            from PIL import Image, ImageTk
            pil_img = Image.open("bandera_tercio_npj.png")
            
            porcentaje = 30  
            
            ancho_orig, alto_orig = pil_img.size
            nuevo_ancho = int(ancho_orig * (porcentaje / 100))
            nuevo_alto = int(alto_orig * (porcentaje / 100))
            
            pil_img_rescalada = pil_img.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)
            logo_img = ImageTk.PhotoImage(pil_img_rescalada)
            
            self.ventana_login.iconphoto(False, logo_img)
            self.root.iconphoto(False, logo_img)
            
            lbl_logo = tk.Label(self.ventana_login, image=logo_img, bg=C_MORADO)
            lbl_logo.image = logo_img  
            lbl_logo.pack(pady=(25, 10))
        except Exception as e:
            pass

        tk.Label(self.ventana_login, text="CONTROL DE ACCESO", font=("Georgia", 18, "bold"), bg=C_MORADO, fg="#d4af37").pack(pady=(0, 25))
        
        tk.Label(self.ventana_login, text="Usuario (Hermandad):", font=("Georgia", 11, "bold"), bg=C_MORADO, fg="#e8d08c").pack(anchor="w", padx=50, pady=(0, 5))
        self.entry_user = tk.Entry(self.ventana_login, font=("Helvetica", 14), bg=C_MORADO_BG, fg="#d4af37", insertbackground="#d4af37", relief="solid", bd=1, justify="center")
        self.entry_user.pack(fill=tk.X, padx=50, ipady=6, pady=(0, 20))
        
        tk.Label(self.ventana_login, text="Contraseña de Licencia:", font=("Georgia", 11, "bold"), bg=C_MORADO, fg="#e8d08c").pack(anchor="w", padx=50, pady=(0, 5))
        self.entry_pass = tk.Entry(self.ventana_login, font=("Helvetica", 14), bg=C_MORADO_BG, fg="#d4af37", insertbackground="#d4af37", show="*", relief="solid", bd=1, justify="center")
        self.entry_pass.pack(fill=tk.X, padx=50, ipady=6, pady=(0, 35))
        
        btn_login = tk.Button(self.ventana_login, text="INICIAR SESIÓN", font=("Georgia", 13, "bold"), bg="#d4af37", fg="#0c0209", activebackground="#b5952f", activeforeground="#000", cursor="hand2", relief="flat", command=self.validar_acceso)
        btn_login.pack(fill=tk.X, padx=50, ipady=8, pady=(0, 20))

    def validar_acceso(self):
        usuario = self.entry_user.get().strip()
        password = self.entry_pass.get().strip()
        
        if not usuario or not password:
            messagebox.showwarning("Atención", "Por favor, rellena ambos campos para verificar tu licencia.")
            return
            
        valido, mensaje = comprobar_licencia(usuario, password)
        
        if valido:
            import json, base64
            pwd_ofuscado = base64.b64encode(password.encode()).decode()
            with open("licencia_local.json", "w") as f:
                json.dump({"usuario": usuario, "password": pwd_ofuscado}, f)
                
            self.ventana_login.destroy()
            self.root.deiconify() 
            messagebox.showinfo("Éxito", "Licencia validada correctamente.\n¡Bienvenido al Gestor Cofrade!")
        else:
            messagebox.showerror("Acceso Denegado", mensaje)

    # --- SISTEMA DE ACTUALIZACIONES ---
    def comprobar_actualizacion_manual(self):
        """Comprueba al hacer clic si hay una versión nueva en la base de datos."""
        try:
            url_firebase = "https://licencias-gestor-cofrade-default-rtdb.europe-west1.firebasedatabase.app/app_config.json"
            resp = requests.get(url_firebase, timeout=5)
            
            if resp.status_code == 200 and resp.json() is not None:
                datos = resp.json()
                
                if "app_config" in datos:
                    datos = datos["app_config"]

                version_nube = datos.get("ultima_version", VERSION_ACTUAL)
                link_descarga = datos.get("link_descarga", "")
                
                if version_nube > VERSION_ACTUAL:
                    respuesta = messagebox.askyesno(
                        "✨ Actualización Disponible", 
                        f"¡Hay una versión nueva ({version_nube}) disponible!\n\n"
                        f"¿Quieres descargarla e instalarla ahora?\n"
                        f"El programa se reiniciará automáticamente."
                    )
                    if respuesta:
                        self.descargar_e_instalar(link_descarga)
                else:
                    messagebox.showinfo("Actualizado", f"Ya tienes la última versión instalada (v{VERSION_ACTUAL}).")
            else:
                messagebox.showwarning("Error", "No se pudo comprobar la versión en la nube en este momento.")
        except Exception as e:
            messagebox.showerror("Error de conexión", f"Comprueba tu conexión a internet.\nDetalle: {e}")

    def descargar_e_instalar(self, url):
        """Descarga el nuevo .exe, ejecuta Actualizador.exe y se cierra."""
        import urllib.request
        import subprocess
        
        # Identificamos el nombre real de este archivo
        if getattr(sys, 'frozen', False):
            exe_actual = os.path.basename(sys.executable)
        else:
            exe_actual = os.path.basename(__file__)

        exe_nuevo = "Gestor_Temp.exe"
        actualizador = "Actualizador.exe"
        
        # Verificar si existe el programa actualizador
        if not os.path.exists(actualizador):
            messagebox.showerror("Falta Actualizador", 
                f"No se ha encontrado el archivo '{actualizador}' en esta misma carpeta.\n"
                "Asegúrate de que está junto al Gestor para poder actualizar.")
            return

        # Mostramos la ventana de espera con diseño moderno
        win_progreso = tk.Toplevel(self.root)
        win_progreso.title("Descargando...")
        win_progreso.geometry("350x120")
        win_progreso.configure(bg=C_MORADO)
        win_progreso.overrideredirect(True)
        
        win_progreso.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (350 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (120 // 2)
        win_progreso.geometry(f"+{x}+{y}")
        
        tk.Label(win_progreso, text="🔄 Descargando actualización...", fg="#d4af37", bg=C_MORADO, font=("Segoe UI", 12, "bold")).pack(pady=(30, 5))
        tk.Label(win_progreso, text="Por favor, no cierres el programa.", fg=C_BLANCO, bg=C_MORADO, font=("Segoe UI", 10)).pack()
        
        # Actualizamos la vista antes de congelar el hilo principal con la descarga
        self.root.update()

        try:
            # Descargamos el archivo a Gestor_Temp.exe
            urllib.request.urlretrieve(url, exe_nuevo)
            
            # Lanzamos el actualizador pasándole qué archivos debe intercambiar
            subprocess.Popen([actualizador, exe_actual, exe_nuevo])
            
            # Cerramos este programa (GestorCofrade.exe) inmediatamente
            self.root.quit()
            sys.exit()
            
        except Exception as e:
            win_progreso.destroy()
            messagebox.showerror("Error en la descarga", f"Hubo un fallo al intentar actualizar:\n{e}")

    # --- UTILIDADES ---
    def guardar_censo(self, datos):
        try:
            with open(CONFIG['archivo_datos'], 'w', encoding='utf-8') as f:
                json.dump(datos, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
            return False

    def abrir_navegador(self, archivo_html):
        from pathlib import Path
        webbrowser.open(Path(archivo_html).resolve().as_uri())

    def mostrar_toast(self, mensaje, duracion=2000, color=None):
        """Notificación flotante en la parte inferior que desaparece sola."""
        if color is None:
            color = C_MORADO
        try:
            toast = tk.Toplevel(self.root)
            toast.overrideredirect(True)
            toast.configure(bg=color)
            toast.attributes('-topmost', True)
            tk.Label(toast, text=mensaje, bg=color, fg=C_BLANCO,
                     font=("Segoe UI", 11, "bold"), padx=22, pady=13).pack()
            toast.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (toast.winfo_width() // 2)
            y = self.root.winfo_y() + self.root.winfo_height() - 80
            toast.geometry(f"+{x}+{y}")
            self.root.after(duracion, lambda: toast.destroy() if toast.winfo_exists() else None)
        except Exception:
            pass

    def mostrar_ficha_guardada(self, nombre, altura, telefono, hombro, miercoles, viernes, editar=False):
        """Popup centrado que muestra los datos guardados y se cierra solo a los 3 segundos."""
        try:
            popup = tk.Toplevel(self.root)
            popup.overrideredirect(True)
            popup.configure(bg=C_MORADO)
            popup.attributes('-topmost', True)

            inner = tk.Frame(popup, bg=C_BLANCO, padx=28, pady=22)
            inner.pack(padx=2, pady=2)

            titulo = "✏️  Costalero actualizado" if editar else "✅  Costalero añadido al censo"
            tk.Label(inner, text=titulo, font=("Segoe UI", 13, "bold"),
                     bg=C_BLANCO, fg=C_MORADO).pack(anchor="w", pady=(0, 12))

            mi_txt = "✅ Sí" if miercoles else "❌ No"
            vi_txt = "✅ Sí" if viernes else "❌ No"
            hombro_txt = hombro if hombro else "Indiferente"

            lineas = (
                f"  👤  {nombre}\n"
                f"  📏  Altura de hombro: {altura} cm\n"
                f"  📱  Teléfono: {telefono or '—'}\n"
                f"  💪  Hombro preferido: {hombro_txt}\n"
                f"  🕯️  Miércoles Santo: {mi_txt}     ✝️  Viernes Santo: {vi_txt}"
            )
            tk.Label(inner, text=lineas, font=("Segoe UI", 11),
                     bg=C_BLANCO, fg="#333", justify="left").pack(anchor="w")
            tk.Label(inner, text="Se cierra automáticamente en 3 segundos…",
                     font=("Segoe UI", 9, "italic"), bg=C_BLANCO, fg="#aaa").pack(anchor="e", pady=(12, 0))

            popup.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (popup.winfo_width() // 2)
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (popup.winfo_height() // 2)
            popup.geometry(f"+{x}+{y}")
            self.root.after(3000, lambda: popup.destroy() if popup.winfo_exists() else None)
        except Exception:
            pass

    def crear_boton_moderno(self, parent, text, bg_color, hover_color, text_color, icon="", command=None, width=30):
        btn_frame = tk.Frame(parent, bg=bg_color, cursor="hand2")
        lbl = tk.Label(btn_frame, text=f"{icon}  {text}" if icon else text, 
                       bg=bg_color, fg=text_color, font=("Segoe UI", 11, "bold"))
        lbl.pack(pady=12, padx=20)
        
        def on_enter(e):
            btn_frame['bg'] = hover_color
            lbl['bg'] = hover_color
        def on_leave(e):
            btn_frame['bg'] = bg_color
            lbl['bg'] = bg_color
            
        btn_frame.bind("<Enter>", on_enter)
        btn_frame.bind("<Leave>", on_leave)
        lbl.bind("<Enter>", on_enter)
        lbl.bind("<Leave>", on_leave)
        
        if command:
            btn_frame.bind("<Button-1>", lambda e: command())
            lbl.bind("<Button-1>", lambda e: command())
            
        return btn_frame

    # --- NAVEGACIÓN Y ANIMACIÓN ---
    def crear_menu_lateral(self):
        lbl_titulo = tk.Label(self.frame_menu, text="GESTOR\nCOSTALEROS", bg=C_MORADO, fg=C_ORO, font=("Georgia", 20, "bold"), pady=30)
        lbl_titulo.pack(fill=tk.X)
        
        opciones = [
            ("Inicio", "🏠"), 
            ("Miércoles Santo", "🕯️"), 
            ("Viernes Santo", "✝️"), 
            ("Vía Crucis", "⛪"), 
            ("Procesión Personalizada", "⚙️"), 
            ("Ensayos", "📋"), 
            ("Calendario", "📅"), 
            ("Censo (Costaleros)", "👥"), 
            ("Publicar / PDF", "📤") 
        ]
        
        for op, icon in opciones:
            btn = self.crear_boton_moderno(self.frame_menu, op, C_MORADO, C_MORADO_HOVER, C_BLANCO, icon=icon, command=lambda nombre=op: self.mostrar_pantalla(nombre))
            btn.pack(fill=tk.X, pady=2, padx=10)
            
        btn_actualizar = self.crear_boton_moderno(self.frame_menu, "Buscar Actualización", "#17517e", "#1f6b9c", C_BLANCO, icon="🔄", command=self.comprobar_actualizacion_manual)
        btn_actualizar.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 5), padx=10)
        
        btn_salir = self.crear_boton_moderno(self.frame_menu, "SALIR DEL SISTEMA", "#ff4757", "#ff6b81", C_BLANCO, icon="❌", command=self.root.quit)
        btn_salir.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 20), padx=10)

    def mostrar_pantalla(self, nombre):
        frame_in = self.frames.get(nombre)
        if not frame_in: return
        if self.current_frame == frame_in: return
        
        if self.current_frame: self.current_frame.place_forget()
        self.current_frame = frame_in
        
        if nombre == "Censo (Costaleros)":
            self.actualizar_tabla_censo()

        def slide(step):
            if step > 0:
                relx = step * 0.01
                frame_in.place(relx=relx, rely=0, relwidth=1, relheight=1)
                self.root.after(10, slide, step - 1)
            else:
                frame_in.place(relx=0.0, rely=0, relwidth=1, relheight=1)
                
        slide(10)

    def crear_pantallas(self):
        self.frames["Inicio"] = self.crear_pantalla_inicio()
        self.frames["Miércoles Santo"] = self.crear_pantalla_procesion("Miércoles Santo", "visualizador_miercoles.html", generar_cuadrillas_miercoles, generar_html_miercoles, True)
        self.frames["Viernes Santo"] = self.crear_pantalla_procesion("Viernes Santo", "visualizador_viernes.html", generar_cuadrillas_viernes, generar_html_viernes, True)
        self.frames["Vía Crucis"] = self.crear_pantalla_viacrucis()
        self.frames["Procesión Personalizada"] = self.crear_pantalla_personalizada()
        self.frames["Ensayos"] = self.crear_pantalla_ensayos()
        self.frames["Calendario"] = self.crear_pantalla_calendario()
        self.frames["Censo (Costaleros)"] = self.crear_pantalla_censo()
        self.frames["Publicar / PDF"] = self.crear_pantalla_pdf()

    # --- PANTALLAS ---
    def crear_pantalla_inicio(self):
        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        
        canvas_fondo = tk.Canvas(f, bg=C_GRIS_FONDO, highlightthickness=0)
        canvas_fondo.place(relwidth=1, relheight=1)

        try:
            from PIL import Image, ImageTk, ImageEnhance, ImageDraw
            import os
            
            ruta_imagen = None
            for ext in ["jpg", "JPG", "jpeg", "JPEG"]:
                candidato = f"cristo_agonia.{ext}"
                if os.path.exists(candidato):
                    ruta_imagen = candidato
                    break
            
            if ruta_imagen and os.path.exists(ruta_imagen):
                img_original = Image.open(ruta_imagen).convert("RGBA")
                
                def redimensionar_fondo(event):
                    if event.width < 100 or event.height < 100: return
                    
                    ratio_img = img_original.width / img_original.height
                    ratio_ventana = event.width / event.height
                    
                    if ratio_ventana > ratio_img:
                        nuevo_ancho = event.width
                        nuevo_alto = int(nuevo_ancho / ratio_img)
                    else:
                        nuevo_alto = event.height
                        nuevo_ancho = int(nuevo_alto * ratio_img)
                        
                    img_redim = img_original.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)
                    
                    enhancer = ImageEnhance.Brightness(img_redim)
                    img_redim = enhancer.enhance(0.5) 
                    
                    capa_transparente = Image.new('RGBA', img_redim.size, (0,0,0,0))
                    draw = ImageDraw.Draw(capa_transparente)
                    
                    card_w, card_h = 800, 250
                    img_cx, img_cy = nuevo_ancho // 2, nuevo_alto // 2
                    
                    x1 = img_cx - card_w // 2
                    y1 = img_cy - card_h // 2
                    x2 = img_cx + card_w // 2
                    y2 = img_cy + card_h // 2
                    
                    draw.rectangle([x1, y1, x2, y2], fill=(255, 255, 255, 180), outline=(212, 175, 55, 255), width=4)
                    
                    img_final = Image.alpha_composite(img_redim, capa_transparente)
                    
                    f.img_tk = ImageTk.PhotoImage(img_final)
                    canvas_fondo.delete("all")
                    canvas_fondo.create_image(event.width/2, event.height/2, image=f.img_tk, anchor="center")
                    
                    cx, cy = event.width // 2, event.height // 2
                    
                    canvas_fondo.create_text(cx, cy - 60, text="Sistema de Gestión Turnos y Procesiones", font=("Cinzel", 24, "bold"), fill=C_MORADO)
                    canvas_fondo.create_text(cx, cy - 20, text="OFS Muy Ilustre Mayordomía de Ntro. Padre Jesús Nazareno", font=("Cinzel", 15), fill="#333333")
                    canvas_fondo.create_text(cx, cy + 10, text="Tercio del Cristo de la Agonía y María Magdalena", font=("Cinzel", 13), fill="#333333")
                    
                    canvas_fondo.create_line(cx - 100, cy + 40, cx + 100, cy + 40, fill=C_ORO, width=2)
                    
                    canvas_fondo.create_text(cx, cy + 70, text="Selecciona un módulo en el menú lateral izquierdo para empezar a trabajar.", font=("Segoe UI", 12), fill="#555555")

                f.bind("<Configure>", redimensionar_fondo)
            else:
                card = tk.Frame(f, bg=C_BLANCO, padx=50, pady=50, highlightbackground=C_ORO, highlightthickness=2)
                card.place(relx=0.5, rely=0.5, anchor="center")
                tk.Label(card, text="Sistema de Gestión Turnos y Procesiones", font=("Georgia", 24, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(pady=(0, 10))
                tk.Label(card, text="OFS Muy Ilustre Mayordomía de Ntro. Padre Jesús Nazareno", font=("Georgia", 15), bg=C_BLANCO, fg="#666").pack(pady=5)
                tk.Label(card, text="Tercio del Cristo de la Agonía y María Magdalena", font=("Georgia", 13), bg=C_BLANCO, fg="#666").pack(pady=5)
                tk.Frame(card, height=2, bg=C_ORO, width=200).pack(pady=20)
                tk.Label(card, text="Selecciona un módulo en el menú lateral izquierdo para empezar a trabajar.", font=("Segoe UI", 12), bg=C_BLANCO, fg="#888").pack(pady=10)

        except Exception as e:
            print("Error cargando imagen de fondo:", e)

        return f

    def crear_pantalla_procesion(self, titulo, html_file, func_generar, func_html, necesita_es_par):
        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        card = tk.Frame(f, bg=C_BLANCO, padx=40, pady=40, highlightbackground="#e0e0e0", highlightthickness=1)
        card.place(relx=0.5, rely=0.45, anchor="center", width=650)
        
        tk.Label(card, text=f"Procesión: {titulo}", font=("Segoe UI", 22, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(anchor="w", pady=(0, 20))
        
        fila_anio = tk.Frame(card, bg=C_BLANCO)
        fila_anio.pack(fill=tk.X, pady=(0, 30))
        tk.Label(fila_anio, text="Año de la procesión:", bg=C_BLANCO, font=("Segoe UI", 12)).pack(side=tk.LEFT)
        entry_anio = tk.Entry(fila_anio, font=("Segoe UI", 14), width=10, bg="#f9f9f9", relief="solid", bd=1)
        entry_anio.insert(0, str(datetime.datetime.now().year))
        entry_anio.pack(side=tk.LEFT, padx=15)

        def generar_nuevo():
            anio = int(entry_anio.get()) if entry_anio.get().isdigit() else datetime.datetime.now().year
            es_par = (anio % 2 == 0)
            master_list = cargar_datos(CONFIG['archivo_datos'])
            
            if "Miércoles" in titulo: master_list = [p for p in master_list if p.get('miercoles_santo', False)]
            else: master_list = [p for p in master_list if p.get('viernes_santo', False)]
                
            if not master_list:
                messagebox.showwarning("Aviso", f"No hay costaleros asignados para el {titulo}.")
                return
            
            if necesita_es_par: datos = func_generar(master_list, es_par)
            else: datos = func_generar(master_list)
                
            func_html(datos, master_list, anio, es_par, CONFIG['peso_trono_kg'], CONFIG['peso_cruz_kg'], CONFIG['limite_peso_persona'])
            self.abrir_navegador(html_file)

            def extraer_ids(obj):
                ids = set()
                if isinstance(obj, dict):
                    if 'id' in obj and obj['id'] != -1 and obj.get('altura', 0) > 0:
                        ids.add(obj['id'])
                    for v in obj.values():
                        ids |= extraer_ids(v)
                elif isinstance(obj, list):
                    for item in obj:
                        ids |= extraer_ids(item)
                return ids

            ids_asignados = extraer_ids(datos)
            sin_asignar = [p for p in master_list if p.get('id', -1) not in ids_asignados and p.get('id', -1) != -1]

            if sin_asignar:
                nombres = "\n".join(f"  • {p['nombre']} ({p['altura']} cm)" for p in sin_asignar)
                messagebox.showwarning(
                    f"⚠️ Costaleros sin asignar — {titulo}",
                    f"Los siguientes costaleros del censo están marcados para el {titulo} pero no han quedado en ningún hueco del cuadrante.\n\nPuedes añadirlos manualmente en el visualizador:\n\n{nombres}"
                )

        def abrir_anterior():
            if os.path.exists(html_file): self.abrir_navegador(html_file)
            else:
                resp = messagebox.askyesno("No encontrado", f"No existe ningún entorno web previo guardado.\n\n¿Quieres generar una plantilla base nueva ahora?")
                if resp: generar_nuevo()

        btn_frame = tk.Frame(card, bg=C_BLANCO)
        btn_frame.pack(fill=tk.X, pady=10)
        
        btn_nuevo = self.crear_boton_moderno(btn_frame, "GENERAR NUEVO CUADRANTE", C_ORO, C_ORO_HOVER, C_TEXTO, command=generar_nuevo)
        btn_nuevo.pack(side=tk.LEFT, padx=(0, 10))
        btn_abrir = self.crear_boton_moderno(btn_frame, "ABRIR CUADRANTE ANTERIOR", "#17517e", "#1f6b9c", C_BLANCO, command=abrir_anterior)
        btn_abrir.pack(side=tk.LEFT)
        tk.Label(card, text="* Generar un nuevo cuadrante sobreescribirá la ultima modificación sin guardar.", font=("Segoe UI", 10, "italic"), bg=C_BLANCO, fg="#888").pack(anchor="w", pady=(30, 0))
        return f

    def crear_pantalla_personalizada(self):
        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        card = tk.Frame(f, bg=C_BLANCO, padx=40, pady=40, highlightbackground="#e0e0e0", highlightthickness=1)
        card.place(relx=0.5, rely=0.45, anchor="center", width=800)
        
        tk.Label(card, text="⚙️ Procesión Personalizada", font=("Segoe UI", 22, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(anchor="w", pady=(0, 20))
        tk.Label(card, text="Configura un entorno de procesión manual. Podrás organizar a los costaleros arrastrándolos desde el menú.", font=("Segoe UI", 11), bg=C_BLANCO, fg="#666").pack(anchor="w", pady=(0, 20))
        
        config_frame = tk.Frame(card, bg=C_BLANCO)
        config_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(config_frame, text="Filtro de Censo:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=10, padx=5)
        var_dia = tk.StringVar(value="Viernes Santo")
        ttk.Combobox(config_frame, textvariable=var_dia, values=["Miércoles Santo", "Viernes Santo", "Todos"], state="readonly", width=18, font=("Segoe UI", 11)).grid(row=0, column=1, sticky="w", pady=10, padx=5)
        
        tk.Label(config_frame, text="Opciones extra:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky="w", pady=10, padx=(30, 5))
        var_cruz = tk.BooleanVar(value=True)
        var_costaleros_7 = tk.BooleanVar(value=False)
        ttk.Checkbutton(config_frame, text="Lleva Cruz Guía", variable=var_cruz).grid(row=0, column=3, sticky="w", pady=5, padx=5)
        ttk.Checkbutton(config_frame, text="7 Costaleros/vara en Cristo", variable=var_costaleros_7).grid(row=0, column=4, sticky="w", pady=5, padx=15)
        
        tk.Label(config_frame, text="Turnos de Trono (Ej: 2):", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="w", pady=10, padx=5)
        entry_turnos_trono = tk.Entry(config_frame, font=("Segoe UI", 12), width=5, relief="solid", bd=1)
        entry_turnos_trono.insert(0, "2")
        entry_turnos_trono.grid(row=1, column=1, sticky="w", pady=10, padx=5)
        
        tk.Label(config_frame, text="Turnos de Cruz (Ej: 3):", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=1, column=2, sticky="w", pady=10, padx=(30, 5))
        entry_turnos_cruz = tk.Entry(config_frame, font=("Segoe UI", 12), width=5, relief="solid", bd=1)
        entry_turnos_cruz.insert(0, "3")
        entry_turnos_cruz.grid(row=1, column=3, sticky="w", pady=10, padx=5, columnspan=2)

        tk.Label(config_frame, text="Tramos Totales:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="w", pady=10, padx=5)
        entry_tramos_totales = tk.Entry(config_frame, font=("Segoe UI", 12), width=5, relief="solid", bd=1)
        entry_tramos_totales.insert(0, "8")
        entry_tramos_totales.grid(row=2, column=1, sticky="w", pady=10, padx=5)

        tk.Label(config_frame, text="IA de Autocompletado:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).grid(row=2, column=2, sticky="w", pady=10, padx=(30, 5))
        var_auto_trono = tk.BooleanVar(value=False)
        var_auto_cruz = tk.BooleanVar(value=False)
        ttk.Checkbutton(config_frame, text="Autocompletar Trono", variable=var_auto_trono).grid(row=2, column=3, sticky="w", pady=5, padx=5)
        ttk.Checkbutton(config_frame, text="Autocompletar Cruces", variable=var_auto_cruz).grid(row=2, column=4, sticky="w", pady=5, padx=15)

        def generar_nuevo():
            dia = var_dia.get()
            lleva_cruz = var_cruz.get()
            costaleros_7 = var_costaleros_7.get()
            auto_trono = var_auto_trono.get()
            auto_cruz = var_auto_cruz.get()
            
            turnos_trono = int(entry_turnos_trono.get()) if entry_turnos_trono.get().isdigit() else 2
            turnos_cruz = int(entry_turnos_cruz.get()) if entry_turnos_cruz.get().isdigit() else 3
            tramos = int(entry_tramos_totales.get()) if entry_tramos_totales.get().isdigit() else 8

            master_list = cargar_datos(CONFIG['archivo_datos'])
            if dia == "Miércoles Santo":
                master_list = [p for p in master_list if p.get('miercoles_santo', False)]
            elif dia == "Viernes Santo":
                master_list = [p for p in master_list if p.get('viernes_santo', False)]

            if not master_list:
                messagebox.showwarning("Aviso", f"No hay costaleros asignados para el filtro: {dia}.")
                return

            datos_gen = generar_datos_personalizados(turnos_trono, turnos_cruz, tramos, lleva_cruz, costaleros_7, auto_trono, auto_cruz, master_list)
            generar_html_personalizado(datos_gen, master_list, lleva_cruz)
            self.abrir_navegador("visualizador_personalizado.html")

        def abrir_anterior():
            if os.path.exists("visualizador_personalizado.html"):
                self.abrir_navegador("visualizador_personalizado.html")
            else:
                messagebox.askyesno("No encontrado", "No existe ninguna procesión personalizada guardada.")

        btn_frame = tk.Frame(card, bg=C_BLANCO)
        btn_frame.pack(fill=tk.X, pady=10)
        btn_generar = self.crear_boton_moderno(btn_frame, "CREAR CUADRANTE MANUAL", C_ORO, C_ORO_HOVER, C_TEXTO, command=generar_nuevo)
        btn_generar.pack(side=tk.LEFT, padx=(0, 10))
        btn_abrir = self.crear_boton_moderno(btn_frame, "ABRIR ANTERIOR", "#17517e", "#1f6b9c", C_BLANCO, command=abrir_anterior)
        btn_abrir.pack(side=tk.LEFT)
        return f

    def crear_pantalla_viacrucis(self):
        import unicodedata
        def norm(s): return unicodedata.normalize('NFD', str(s)).encode('ascii','ignore').decode().lower()

        self._vc_sort_col = "Nombre"
        self._vc_sort_rev = False

        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        
        tk.Label(f, text="⛪ Gestor de Vía Crucis", font=("Segoe UI", 22, "bold"), bg=C_GRIS_FONDO, fg=C_MORADO).pack(anchor="w", padx=30, pady=(30, 10))
        
        card = tk.Frame(f, bg=C_BLANCO, padx=20, pady=20, highlightbackground="#e0e0e0", highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 30))
        
        config_frame = tk.Frame(card, bg=C_BLANCO)
        config_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(config_frame, text="Número de Tramos/Relevos:", bg=C_BLANCO, font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        entry_tramos = tk.Entry(config_frame, font=("Segoe UI", 12), width=5, relief="solid", bd=1)
        entry_tramos.insert(0, "4")
        entry_tramos.pack(side=tk.LEFT, padx=10)
        
        var_auto = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, text="Autocompletar turnos (por altura)", variable=var_auto).pack(side=tk.LEFT, padx=20)

        search_frame = tk.Frame(card, bg=C_BLANCO)
        search_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(search_frame, text="🔍 Buscar:", bg=C_BLANCO, font=("Segoe UI", 11)).pack(side=tk.LEFT, padx=(0, 5))
        entry_busq_vc = tk.Entry(search_frame, font=("Segoe UI", 11), width=28, relief="solid", bd=1)
        entry_busq_vc.pack(side=tk.LEFT)

        lbl_cont_vc = tk.Label(search_frame, text="", bg=C_BLANCO, font=("Segoe UI", 10), fg="#555")
        lbl_cont_vc.pack(side=tk.RIGHT)

        tk.Label(card, text="☑ Selecciona los costaleros que asistirán haciendo clic en cualquier parte de su fila:", bg=C_BLANCO, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(5, 5))
        
        frame_tabla = tk.Frame(card, bg=C_BLANCO)
        frame_tabla.pack(fill=tk.BOTH, expand=True)
        
        columnas = ("Sel", "Nombre", "Altura", "Hombro", "ID")
        self.tree_viacrucis = ttk.Treeview(frame_tabla, columns=columnas, show="headings", selectmode="none")

        def cmd_sort_vc(col):
            if self._vc_sort_col == col:
                self._vc_sort_rev = not self._vc_sort_rev
            else:
                self._vc_sort_col = col
                self._vc_sort_rev = False
            refrescar_vc()

        self.tree_viacrucis.heading("Sel",    text="✔")
        self.tree_viacrucis.heading("Nombre", text="Nombre",   command=lambda: cmd_sort_vc("Nombre"))
        self.tree_viacrucis.heading("Altura", text="Altura",   command=lambda: cmd_sort_vc("Altura"))
        self.tree_viacrucis.heading("Hombro", text="Hombro")
        self.tree_viacrucis.heading("ID",     text="ID")
        
        self.tree_viacrucis.column("Sel",    width=50,  anchor="center")
        self.tree_viacrucis.column("Nombre", width=250)
        self.tree_viacrucis.column("Altura", width=100, anchor="center")
        self.tree_viacrucis.column("Hombro", width=120, anchor="center")
        self.tree_viacrucis.column("ID",     width=0,   stretch=tk.NO)
        
        self.tree_viacrucis.tag_configure("selected",   background=C_ORO,    foreground=C_TEXTO)
        self.tree_viacrucis.tag_configure("unselected", background=C_BLANCO, foreground=C_TEXTO)

        scroll = ttk.Scrollbar(frame_tabla, orient=tk.VERTICAL, command=self.tree_viacrucis.yview)
        self.tree_viacrucis.configure(yscroll=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_viacrucis.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._vc_seleccionados = set() 

        def toggle_selection(event):
            region = self.tree_viacrucis.identify("region", event.x, event.y)
            if region == "cell":
                item = self.tree_viacrucis.identify_row(event.y)
                if item:
                    pid = int(self.tree_viacrucis.item(item, "values")[4])
                    if pid in self._vc_seleccionados:
                        self._vc_seleccionados.discard(pid)
                        self.tree_viacrucis.item(item, tags=("unselected",))
                    else:
                        self._vc_seleccionados.add(pid)
                        self.tree_viacrucis.item(item, tags=("selected",))
                    _actualizar_contador_vc()

        self.tree_viacrucis.bind("<ButtonRelease-1>", toggle_selection)

        def _actualizar_contador_vc():
            total = len(self.tree_viacrucis.get_children())
            sel   = len(self._vc_seleccionados)
            lbl_cont_vc.config(text=f"{sel} seleccionados de {total}")

        def refrescar_vc(event=None):
            filtro = norm(entry_busq_vc.get())
            datos = cargar_datos(CONFIG['archivo_datos'])

            if filtro:
                datos = [p for p in datos if
                         filtro in norm(p.get('nombre', ''))
                         or str(p.get('altura', '')) == filtro]

            col = self._vc_sort_col
            rev = self._vc_sort_rev
            etiquetas = {"Nombre": "Nombre", "Altura": "Altura"}
            for c, txt in etiquetas.items():
                indicador = (" ▼" if rev else " ▲") if c == col else ""
                self.tree_viacrucis.heading(c, text=txt + indicador)

            if col == "Nombre": datos.sort(key=lambda x: norm(x.get('nombre', '')), reverse=rev)
            elif col == "Altura": datos.sort(key=lambda x: x.get('altura', 0), reverse=rev)

            for item in self.tree_viacrucis.get_children():
                self.tree_viacrucis.delete(item)

            for d in datos:
                hombro = str(d.get('pref_hombro', '')).capitalize()
                if not hombro or hombro.lower() in ("indiferente", ""):
                    hombro = "-"
                pid = d.get('id', '')
                check = "☑" if pid in self._vc_seleccionados else "☐"
                tag   = "selected" if pid in self._vc_seleccionados else "unselected"
                self.tree_viacrucis.insert("", "end",
                    values=(check, d.get('nombre',''), f"{d.get('altura','')} cm", hombro, pid),
                    tags=(tag,))
            _actualizar_contador_vc()

        entry_busq_vc.bind("<KeyRelease>", refrescar_vc)
        refrescar_vc()

        def generar():
            datos_todos = cargar_datos(CONFIG['archivo_datos'])
            ids_sel = list(self._vc_seleccionados)
                    
            if not ids_sel:
                if not messagebox.askyesno("Atención", "No has seleccionado a ningún costalero.\n\n¿Quieres generar los tramos del Vía Crucis completamente vacíos para rellenarlos a mano?"):
                    return
            
            lista_seleccionados = [d for d in datos_todos if d.get('id') in ids_sel]
            tramos = int(entry_tramos.get()) if entry_tramos.get().isdigit() else 4
            datos_gen = generar_datos_viacrucis(lista_seleccionados, tramos, var_auto.get() and len(ids_sel) > 0)
            generar_html_viacrucis(datos_gen, datos_todos, ids_sel)
            self.abrir_navegador("visualizador_viacrucis.html")
            
        def abrir_anterior():
            if os.path.exists("visualizador_viacrucis.html"): self.abrir_navegador("visualizador_viacrucis.html")
            else: messagebox.askyesno("No encontrado", "No existe ningún Vía Crucis anterior guardado.")

        btn_frame = tk.Frame(card, bg=C_BLANCO)
        btn_frame.pack(fill=tk.X, pady=15)
        btn_generar = self.crear_boton_moderno(btn_frame, "GENERAR VÍA CRUCIS", C_ORO, C_ORO_HOVER, C_TEXTO, command=generar)
        btn_generar.pack(side=tk.LEFT, padx=(0, 10))
        btn_abrir = self.crear_boton_moderno(btn_frame, "ABRIR ANTERIOR", "#17517e", "#1f6b9c", C_BLANCO, command=abrir_anterior)
        btn_abrir.pack(side=tk.LEFT)
        return f

    def crear_pantalla_ensayos(self):
        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        card = tk.Frame(f, bg=C_BLANCO, padx=40, pady=40, highlightbackground="#e0e0e0", highlightthickness=1)
        card.place(relx=0.5, rely=0.45, anchor="center", width=650)
        
        tk.Label(card, text="Organizador de Ensayos", font=("Segoe UI", 22, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(anchor="w", pady=(0, 20))
        
        fila_turnos = tk.Frame(card, bg=C_BLANCO)
        fila_turnos.pack(fill=tk.X, pady=(0, 30))
        tk.Label(fila_turnos, text="Número de Turnos (Ej: 2 para 72 costaleros):", bg=C_BLANCO, font=("Segoe UI", 12)).pack(side=tk.LEFT)
        entry_turnos = tk.Entry(fila_turnos, font=("Segoe UI", 14), width=8, bg="#f9f9f9", relief="solid", bd=1)
        entry_turnos.insert(0, "2")
        entry_turnos.pack(side=tk.LEFT, padx=15)
        
        def generar_nuevo():
            turnos = int(entry_turnos.get()) if entry_turnos.get().isdigit() else 2
            master_list = cargar_datos(CONFIG['archivo_datos'])
            generar_html_ensayo(turnos, master_list, CONFIG['peso_trono_kg'], CONFIG['limite_peso_persona'])
            self.abrir_navegador("visualizador_ensayos.html")
            
        def abrir_anterior():
            if os.path.exists("visualizador_ensayos.html"): self.abrir_navegador("visualizador_ensayos.html")
            else:
                resp = messagebox.askyesno("No encontrado", "No existe ningún entorno de ensayo.\n\n¿Quieres crear el entorno ahora?")
                if resp: generar_nuevo()

        btn_frame = tk.Frame(card, bg=C_BLANCO)
        btn_frame.pack(fill=tk.X, pady=10)
        btn_nuevo = self.crear_boton_moderno(btn_frame, "NUEVO ENTORNO", C_ORO, C_ORO_HOVER, C_TEXTO, command=generar_nuevo)
        btn_nuevo.pack(side=tk.LEFT, padx=(0, 10))
        btn_abrir = self.crear_boton_moderno(btn_frame, "ABRIR ANTERIOR", "#17517e", "#1f6b9c", C_BLANCO, command=abrir_anterior)
        btn_abrir.pack(side=tk.LEFT)
        return f

    def crear_pantalla_calendario(self):
        f = tk.Frame(self.frame_main, bg="#f4f6f8", padx=30, pady=30)
        
        tk.Label(f, text="📅 Calendario de Ensayos y Citas", font=("Segoe UI", 22, "bold"), bg="#f4f6f8", fg=C_MORADO).pack(anchor="w")
        
        toolbar = tk.Frame(f, bg="#f4f6f8")
        toolbar.pack(fill=tk.X, pady=15)
        
        btn_nuevo = self.crear_boton_moderno(toolbar, "➕ Añadir Cita", C_ORO, C_ORO_HOVER, C_TEXTO, command=lambda: self.abrir_formulario_evento())
        btn_nuevo.pack(side=tk.LEFT, padx=(0, 10))
        btn_editar = self.crear_boton_moderno(toolbar, "✏️ Editar", "#17517e", "#1f6b9c", C_BLANCO, command=lambda: self.abrir_formulario_evento(editar=True))
        btn_editar.pack(side=tk.LEFT, padx=10)
        btn_borrar = self.crear_boton_moderno(toolbar, "❌ Borrar", "#ff4757", "#ff6b81", C_BLANCO, command=self.borrar_evento)
        btn_borrar.pack(side=tk.LEFT, padx=10)
        
        def exportar_html():
            exito, archivo = generar_html_calendario(self.lista_eventos)
            if exito: self.abrir_navegador(archivo)
                
        btn_pdf = self.crear_boton_moderno(toolbar, "📤 Exportar PDF", C_MORADO, C_MORADO_HOVER, C_BLANCO, command=exportar_html)
        btn_pdf.pack(side=tk.RIGHT)

        frame_tabla = tk.Frame(f, bg=C_BLANCO, highlightbackground="#ccc", highlightthickness=1)
        frame_tabla.pack(fill=tk.BOTH, expand=True)
        
        columnas = ("id", "fecha", "hora", "motivo", "lugar", "indicaciones")
        self.tabla_calendario = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=15)
        
        self.tabla_calendario.heading("id", text="ID") 
        self.tabla_calendario.heading("fecha", text="Día")
        self.tabla_calendario.heading("hora", text="Hora")
        self.tabla_calendario.heading("motivo", text="Motivo / Cita")
        self.tabla_calendario.heading("lugar", text="Lugar")
        self.tabla_calendario.heading("indicaciones", text="Indicaciones")
        
        self.tabla_calendario.column("id", width=0, stretch=tk.NO)
        self.tabla_calendario.column("fecha", width=120, anchor="center")
        self.tabla_calendario.column("hora", width=80, anchor="center")
        self.tabla_calendario.column("motivo", width=200)
        self.tabla_calendario.column("lugar", width=150)
        self.tabla_calendario.column("indicaciones", width=250)
        
        scrollbar = ttk.Scrollbar(frame_tabla, orient=tk.VERTICAL, command=self.tabla_calendario.yview)
        self.tabla_calendario.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tabla_calendario.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.refrescar_tabla_calendario()
        return f

    def refrescar_tabla_calendario(self):
        for row in self.tabla_calendario.get_children():
            self.tabla_calendario.delete(row)
        self.lista_eventos = cargar_eventos()
        for idx, ev in enumerate(self.lista_eventos):
            self.tabla_calendario.insert("", "end", values=(idx, ev.get("fecha",""), ev.get("hora",""), ev.get("motivo",""), ev.get("lugar",""), ev.get("indicaciones","")))

    def borrar_evento(self):
        seleccionado = self.tabla_calendario.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Selecciona una cita de la tabla para borrarla.")
            return
            
        motivo = self.tabla_calendario.item(seleccionado[0])['values'][3]
        
        if messagebox.askyesno("⚠️ Confirmar Borrado", f"¿Estás seguro de eliminar el evento:\n'{motivo}'?"):
            idx = int(self.tabla_calendario.item(seleccionado[0])['values'][0])
            del self.lista_eventos[idx]
            guardar_eventos(self.lista_eventos)
            self.refrescar_tabla_calendario()

    def abrir_formulario_evento(self, editar=False):
        evento = None
        idx_editar = -1
        
        if editar:
            seleccionado = self.tabla_calendario.selection()
            if not seleccionado:
                messagebox.showwarning("Atención", "Selecciona una cita de la tabla para editar.")
                return
            idx_editar = int(self.tabla_calendario.item(seleccionado[0])['values'][0])
            evento = self.lista_eventos[idx_editar]

        top = tk.Toplevel(self.root)
        top.title("Editar Evento" if editar else "Nueva Cita / Ensayo")
        top.geometry("450x620")
        top.configure(bg=C_BLANCO)
        
        top.resizable(True, True)
        top.transient(self.root) 
        top.grab_set()

        top.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (450 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (620 // 2)
        top.geometry(f"+{x}+{y}")

        tk.Label(top, text="📅 Detalles de la Convocatoria", font=("Segoe UI", 16, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(pady=(20, 15))
        
        form = tk.Frame(top, bg=C_BLANCO)
        form.pack(padx=30, fill=tk.BOTH, expand=True)

        fila_fechahora = tk.Frame(form, bg=C_BLANCO)
        fila_fechahora.pack(fill=tk.X, pady=(0, 15))

        bloque_fecha = tk.Frame(fila_fechahora, bg=C_BLANCO)
        bloque_fecha.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        tk.Label(bloque_fecha, text="Día del Evento:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 5))
        cal_fecha = DateEntry(bloque_fecha, background=C_MORADO, foreground='white', borderwidth=2, font=("Segoe UI", 12), date_pattern='dd/MM/yyyy', cursor="hand2")
        cal_fecha.pack(fill=tk.X)
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        if evento and evento.get('fecha'):
            try:
                partes = evento['fecha'].split(' de ')
                dia = int(partes[0])
                mes = meses.index(partes[1]) + 1
                anio_actual = datetime.datetime.now().year
                fecha_obj = datetime.date(anio_actual, mes, dia)
                cal_fecha.set_date(fecha_obj)
            except: pass

        bloque_hora = tk.Frame(fila_fechahora, bg=C_BLANCO)
        bloque_hora.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        tk.Label(bloque_hora, text="Hora:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 5))
        
        caja_selectores = tk.Frame(bloque_hora, bg=C_BLANCO)
        caja_selectores.pack(fill=tk.X)
        
        horas = [str(i).zfill(2) for i in range(24)]
        minutos = ["00", "15", "30", "45"]
        var_hh = tk.StringVar(value="20")
        var_mm = tk.StringVar(value="00")
        
        if evento:
            h_parts = evento.get('hora','').replace('h','').split(':')
            if len(h_parts) == 2:
                var_hh.set(h_parts[0])
                var_mm.set(h_parts[1])

        ttk.Combobox(caja_selectores, textvariable=var_hh, values=horas, width=4, font=("Segoe UI", 12), state="normal").pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Label(caja_selectores, text=":", bg=C_BLANCO, font=("Segoe UI", 12, "bold")).pack(side=tk.LEFT, padx=2)
        ttk.Combobox(caja_selectores, textvariable=var_mm, values=minutos, width=4, font=("Segoe UI", 12), state="normal").pack(side=tk.LEFT, expand=True, fill=tk.X)

        tk.Label(form, text="Motivo (Puedes elegir o escribir uno):", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 5))
        var_motivo = tk.StringVar(value=evento.get('motivo', 'Ensayo') if evento else 'Ensayo')
        combo_motivo = ttk.Combobox(form, textvariable=var_motivo, values=["Ensayo", "Reunión de Costaleros", "Mudá del Trono", "Misa de Hermandad"], font=("Segoe UI", 12))
        combo_motivo.pack(fill=tk.X, pady=(0, 15))

        tk.Label(form, text="Lugar:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 5))
        var_lugar = tk.StringVar(value=evento.get('lugar', 'San Francisco') if evento else 'San Francisco')
        combo_lugar = ttk.Combobox(form, textvariable=var_lugar, values=["San Francisco", "Santuario de Monserrate", "Casa de Hermandad", "As de Oros"], font=("Segoe UI", 12))
        combo_lugar.pack(fill=tk.X, pady=(0, 15))

        tk.Label(form, text="Indicaciones (Opcional):", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 5))
        text_ind = tk.Text(form, font=("Segoe UI", 11), height=4, relief="solid", bd=1, wrap=tk.WORD)
        text_ind.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        placeholder = "Ej: Venir con calzado oscuro, ropa cómoda..."
        
        if evento and evento.get('indicaciones'):
            text_ind.insert("1.0", evento.get('indicaciones'))
            text_ind.config(fg="black")
        else:
            text_ind.insert("1.0", placeholder)
            text_ind.config(fg="grey")

        def on_focus_in(e):
            contenido = text_ind.get("1.0", "end-1c")
            if contenido == placeholder:
                text_ind.delete("1.0", tk.END)
                text_ind.config(fg="black")
                
        def on_focus_out(e):
            contenido = text_ind.get("1.0", "end-1c").strip()
            if not contenido:
                text_ind.insert("1.0", placeholder)
                text_ind.config(fg="grey")

        text_ind.bind("<FocusIn>", on_focus_in)
        text_ind.bind("<FocusOut>", on_focus_out)

        def guardar():
            mtv = var_motivo.get().strip()
            if not mtv:
                messagebox.showwarning("Atención", "Debes especificar un motivo.")
                return
            
            ind_final = text_ind.get("1.0", "end-1c").strip()
            if ind_final == placeholder:
                ind_final = ""

            fecha_seleccionada = cal_fecha.get_date()
            dia_str = str(fecha_seleccionada.day).zfill(2)
            mes_str = meses[fecha_seleccionada.month - 1]
            fecha_formateada = f"{dia_str} de {mes_str}"

            datos_evento = {
                "fecha": fecha_formateada,
                "hora": f"{var_hh.get()}:{var_mm.get()}h",
                "motivo": mtv,
                "lugar": var_lugar.get().strip(),
                "indicaciones": ind_final
            }

            if editar:
                self.lista_eventos[idx_editar] = datos_evento
            else:
                self.lista_eventos.append(datos_evento)
                
            guardar_eventos(self.lista_eventos)
            self.refrescar_tabla_calendario()
            top.destroy()
            accion = "actualizado" if editar else "añadido"
            self.mostrar_toast(f"✅  Evento {accion} correctamente")

        top.bind("<Escape>", lambda e: top.destroy())

        btn_guardar = tk.Button(top, text="💾 GUARDAR EVENTO", bg=C_MORADO, fg=C_BLANCO, font=("Segoe UI", 12, "bold"), bd=0, cursor="hand2", command=guardar)
        btn_guardar.pack(fill=tk.X, padx=30, pady=(0, 20), ipady=8)


    def crear_pantalla_pdf(self):
        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO)
        card = tk.Frame(f, bg=C_BLANCO, padx=40, pady=40, highlightbackground="#e0e0e0", highlightthickness=1)
        card.place(relx=0.5, rely=0.45, anchor="center", width=700)
        
        tk.Label(card, text="📤 Centro de Exportación y Publicación", font=("Segoe UI", 22, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(anchor="w", pady=(0, 10))
        tk.Label(card, text="Genera el acta oficial en PDF o publica el cuadrante en el Portal Web de la Agonía.", font=("Segoe UI", 12), bg=C_BLANCO, fg="#666").pack(anchor="w", pady=(0, 20))
        
        fila_anio = tk.Frame(card, bg=C_BLANCO)
        fila_anio.pack(fill=tk.X, pady=(0, 20))
        tk.Label(fila_anio, text="Año de la procesión a exportar:", bg=C_BLANCO, font=("Segoe UI", 12, "bold")).pack(side=tk.LEFT)
        entry_anio = tk.Entry(fila_anio, font=("Segoe UI", 14), width=10, bg="#f9f9f9", relief="solid", bd=1)
        entry_anio.insert(0, str(datetime.datetime.now().year))
        entry_anio.pack(side=tk.LEFT, padx=15)

        var_tipo = tk.StringVar(value="Viernes Santo")
        style = ttk.Style()
        style.configure("TRadiobutton", background=C_BLANCO, font=("Segoe UI", 12))
        
        ttk.Radiobutton(card, text="Miércoles Santo", variable=var_tipo, value="Miércoles Santo").pack(anchor="w", pady=5)
        ttk.Radiobutton(card, text="Viernes Santo", variable=var_tipo, value="Viernes Santo").pack(anchor="w", pady=5)
        
        estado_exportacion = {
            "archivo_ruta": None,
            "datos_json": None,
            "revisado": False
        }

        panel_pub = tk.Frame(card, bg="#f4f6f8", bd=1, relief="solid", padx=20, pady=20)
        
        lbl_info_archivo = tk.Label(panel_pub, text="", font=("Segoe UI", 11, "bold"), bg="#f4f6f8", fg=C_MORADO)
        lbl_info_archivo.pack(anchor="w", pady=(0, 15))

        btn_vista_previa = tk.Button(panel_pub, text="👁️ 1. VISTA PREVIA E IMPRIMIR PDF", font=("Segoe UI", 12, "bold"), bg="#17517e", fg="white", cursor="hand2")
        btn_vista_previa.pack(fill=tk.X, pady=5, ipady=5)

        btn_publicar = tk.Button(panel_pub, text="🌐 2. PUBLICAR EN EL PORTAL WEB", font=("Segoe UI", 12, "bold"), bg="#27ae60", fg="white", cursor="hand2", state="disabled")
        btn_publicar.pack(fill=tk.X, pady=5, ipady=5)
        
        tk.Label(panel_pub, text="* Es obligatorio revisar la vista previa (PDF) antes de poder publicar en internet.", font=("Segoe UI", 9, "italic"), bg="#f4f6f8", fg="#666").pack(anchor="w", pady=(0,15))

        btn_despublicar = tk.Button(panel_pub, text="🗑️ DESPUBLICAR Y OCULTAR DE LA WEB", font=("Segoe UI", 10, "bold"), bg="#ff4757", fg="white", cursor="hand2")
        btn_despublicar.pack(fill=tk.X, pady=(10,0), ipady=4)

        def cargar_archivo_cuadrante():
            archivo = filedialog.askopenfilename(title="Selecciona el cuadrante final (.json)", filetypes=[("Archivos JSON", "*.json")])
            if archivo:
                try:
                    with open(archivo, 'r', encoding='utf-8') as f:
                        datos = json.load(f)
                    
                    estado_exportacion["archivo_ruta"] = archivo
                    estado_exportacion["datos_json"] = datos
                    estado_exportacion["revisado"] = False
                    
                    t_proc = datos.get("tipo_procesion", "")
                    if var_tipo.get() == "Viernes Santo" and t_proc != "viernes_santo":
                        messagebox.showwarning("Atención", "Has marcado Viernes Santo pero has cargado un archivo del Miércoles Santo.\nEl PDF podría generarse mal.")
                    elif var_tipo.get() == "Miércoles Santo" and t_proc != "miercoles_santo":
                        messagebox.showwarning("Atención", "Has marcado Miércoles Santo pero has cargado un archivo del Viernes Santo.\nEl PDF podría generarse mal.")
                    
                    btn_publicar.config(state="disabled") 
                    lbl_info_archivo.config(text=f"✅ Archivo cargado y listo:\n{os.path.basename(archivo)}")
                    panel_pub.pack(fill=tk.X, pady=15)
                    
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

        def hacer_vista_previa():
            if not estado_exportacion["archivo_ruta"]:
                return
            
            anio = int(entry_anio.get()) if entry_anio.get().isdigit() else datetime.datetime.now().year
            exito, msg = crear_html_informe(var_tipo.get(), estado_exportacion["archivo_ruta"], anio)
            
            if exito:
                self.abrir_navegador(msg)
                estado_exportacion["revisado"] = True
                btn_publicar.config(state="normal") 
                messagebox.showinfo("Revisión Necesaria", "Se ha abierto el borrador del PDF en tu navegador.\n\nRevísalo a fondo. Si ves que está todo perfecto, vuelve aquí y pulsa el botón verde de 'PUBLICAR EN EL PORTAL WEB'.")
            else:
                messagebox.showerror("Error", f"Error al generar la vista previa: {msg}")

        def publicar():
            if not estado_exportacion["revisado"]:
                messagebox.showwarning("Aviso de Seguridad", "Por seguridad, el sistema exige que previsualices el PDF antes de mandarlo a internet.")
                return
                
            datos_subir = estado_exportacion["datos_json"]
            tipo_procesion_nube = "miercoles_santo" if var_tipo.get() == "Miércoles Santo" else "viernes_santo"
            url_firebase = f"https://licencias-gestor-cofrade-default-rtdb.europe-west1.firebasedatabase.app/cuadrantes/agonia/{tipo_procesion_nube}.json"
            
            try:
                btn_publicar.config(text="⏳ COMPROBANDO NUBE...", state="disabled")
                self.root.update()
                
                resp_get = requests.get(url_firebase, timeout=5)
                
                if resp_get.status_code == 200 and resp_get.json() is not None:
                    confirmacion = messagebox.askyesno(
                        "⚠️ ¡ATENCIÓN: SOBREESCRITURA!", 
                        f"El sistema ha detectado que YA EXISTE un cuadrante del {var_tipo.get()} publicado en el Portal Web.\n\n¿Estás completamente seguro de que deseas SOBREESCRIBIRLO con la nueva versión que acabas de revisar?"
                    )
                    if not confirmacion:
                        btn_publicar.config(text="🌐 2. PUBLICAR EN EL PORTAL WEB", state="normal")
                        return
                else:
                    confirmacion = messagebox.askyesno(
                        "Confirmar Publicación", 
                        f"Vas a hacer público en internet el cuadrante del {var_tipo.get()}.\nLos costaleros que tengan el enlace podrán verlo.\n\n¿Deseas continuar?"
                    )
                    if not confirmacion:
                        btn_publicar.config(text="🌐 2. PUBLICAR EN EL PORTAL WEB", state="normal")
                        return

                btn_publicar.config(text="⏳ SUBIENDO A LA NUBE...")
                self.root.update()
                
                resp_put = requests.put(url_firebase, json=datos_subir, timeout=10)
                
                if resp_put.status_code == 200:
                    messagebox.showinfo("✅ Publicación Exitosa", f"¡El cuadrante del {var_tipo.get()} ya está publicado en internet de forma segura!\n\nYa puedes compartir el enlace del Portal Web con los costaleros.")
                else:
                    messagebox.showerror("Error del Servidor", f"Ocurrió un problema en la nube al subir los datos.\nCódigo de error: {resp_put.status_code}")
                    
            except Exception as e:
                messagebox.showerror("Error de Conexión", f"No se ha podido conectar con el Portal Web.\nRevisa tu conexión a internet.\n\nDetalle técnico: {str(e)}")
            finally:
                btn_publicar.config(text="🌐 2. PUBLICAR EN EL PORTAL WEB", state="normal")

        def despublicar():
            tipo_procesion_nube = "miercoles_santo" if var_tipo.get() == "Miércoles Santo" else "viernes_santo"
            url_firebase = f"https://licencias-gestor-cofrade-default-rtdb.europe-west1.firebasedatabase.app/cuadrantes/agonia/{tipo_procesion_nube}.json"
            
            confirmacion = messagebox.askyesno(
                "⚠️ Confirmar Despublicación", 
                f"Estás a punto de ELIMINAR DE INTERNET el cuadrante público del {var_tipo.get()}.\n\nLos costaleros dejarán de tener acceso al mismo y verán un aviso de 'Cuadrante no disponible'.\n\n¿Estás completamente seguro?"
            )
            
            if not confirmacion: 
                return
            
            try:
                resp_del = requests.delete(url_firebase, timeout=5)
                if resp_del.status_code == 200:
                    messagebox.showinfo("🗑️ Despublicación Correcta", f"El cuadrante del {var_tipo.get()} se ha eliminado correctamente de la web.\n\nLa información ya es privada de nuevo.")
                else:
                    messagebox.showerror("Error", f"No se pudo eliminar el archivo en la nube: {resp_del.status_code}")
            except Exception as e:
                messagebox.showerror("Error de Conexión", f"No se pudo conectar a la nube.\n{str(e)}")

        btn_vista_previa.config(command=hacer_vista_previa)
        btn_publicar.config(command=publicar)
        btn_despublicar.config(command=despublicar)

        btn_cargar = self.crear_boton_moderno(card, "📂 1º SELECCIONAR CUADRANTE (.JSON)", C_MORADO, C_MORADO_HOVER, C_BLANCO, command=cargar_archivo_cuadrante)
        btn_cargar.pack(anchor="w", fill=tk.X)
        return f

    def exportar_censo_pdf(self):
        """Genera un HTML del censo con botón de descarga PDF y lo abre en el navegador."""
        import unicodedata
        datos = cargar_datos(CONFIG['archivo_datos'])
        if not datos:
            messagebox.showwarning("Aviso", "El censo está vacío.")
            return

        datos.sort(key=lambda x: unicodedata.normalize('NFD', x.get('nombre', '')).encode('ascii','ignore').decode().lower())

        filas_html = ""
        for p in datos:
            mi = "✅" if p.get('miercoles_santo') else "❌"
            vi = "✅" if p.get('viernes_santo') else "❌"
            hombro = p.get('pref_hombro', '') or "Indiferente"
            tel = p.get('telefono', '') or "—"
            filas_html += f"""
            <tr>
                <td style="text-align:center">{p['id']}</td>
                <td>{p['nombre']}</td>
                <td style="text-align:center">{p['altura']} cm</td>
                <td style="text-align:center">{tel}</td>
                <td style="text-align:center">{hombro.capitalize()}</td>
                <td style="text-align:center">{mi}</td>
                <td style="text-align:center">{vi}</td>
            </tr>"""

        n_mi = sum(1 for p in datos if p.get('miercoles_santo'))
        n_vi = sum(1 for p in datos if p.get('viernes_santo'))
        fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y")

        html = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Censo de Costaleros</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@600&family=Open+Sans:wght@400;600&display=swap');
        body {{ font-family: 'Open Sans', sans-serif; background: #f4f6f8; margin: 0; padding: 20px; }}
        .btn-pdf {{ background: #4F1243; color: #fff; padding: 12px 28px; border: none; border-radius: 5px; font-size: 14px; font-weight: bold; cursor: pointer; text-transform: uppercase; margin-bottom: 20px; display: block; }}
        .container {{ max-width: 900px; margin: 0 auto; background: #fff; padding: 40px; box-shadow: 0 0 15px rgba(0,0,0,.1); border-top: 8px solid #4F1243; }}
        h1 {{ color: #4F1243; font-family: 'Cinzel', serif; font-size: 18px; text-transform: uppercase; margin: 0 0 4px; }}
        h2 {{ color: #b5952f; font-size: 14px; margin: 0 0 6px; letter-spacing: 1px; }}
        .fecha {{ color: #888; font-size: 12px; margin-bottom: 20px; }}
        .resumen {{ display: flex; gap: 20px; margin-bottom: 20px; }}
        .stat {{ background: #f4f6f8; padding: 10px 20px; border-radius: 6px; border-left: 4px solid #4F1243; font-size: 13px; }}
        .stat b {{ font-size: 22px; display: block; color: #4F1243; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
        th {{ background: #d4af37; color: #333; padding: 10px; text-align: left; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #eee; }}
        tr:nth-child(even) td {{ background: #fafafa; }}
        @media print {{ .btn-pdf {{ display: none !important; }} body {{ background: #fff; padding: 0; }} .container {{ box-shadow: none; border-top: none; }} }}
    </style>
</head>
<body>
    <button class="btn-pdf" onclick="generarPDF()">📥 DESCARGAR PDF</button>
    <div class="container" id="censo-content">
        <h1>OFS Muy Ilustre Mayordomía de Nuestro Padre Jesús Nazareno</h1>
        <h2>Censo Oficial de Costaleros — Tercio del Cristo de la Agonía</h2>
        <p class="fecha">Generado el {fecha_hoy}</p>
        <div class="resumen">
            <div class="stat"><b>{len(datos)}</b> Costaleros en censo</div>
            <div class="stat"><b>{n_mi}</b> 🕯️ Miércoles Santo</div>
            <div class="stat"><b>{n_vi}</b> ✝️ Viernes Santo</div>
        </div>
        <table>
            <thead><tr><th>ID</th><th>Nombre</th><th>Altura</th><th>Teléfono</th><th>Hombro</th><th>M. Santo</th><th>V. Santo</th></tr></thead>
            <tbody>{filas_html}</tbody>
        </table>
    </div>
    <script>
        function generarPDF() {{
            const opt = {{ margin: 10, filename: 'Censo_Agonia.pdf', image: {{ type: 'jpeg', quality: 0.98 }}, html2canvas: {{ scale: 2 }}, jsPDF: {{ unit: 'mm', format: 'a4', orientation: 'landscape' }} }};
            html2pdf().set(opt).from(document.getElementById('censo-content')).save();
        }}
    </script>
</body>
</html>"""

        with open("censo_exportado.html", "w", encoding="utf-8") as f:
            f.write(html)
        self.abrir_navegador("censo_exportado.html")
        self._ultimo_formato = "PDF"

    def exportar_censo_excel(self):
        """Exporta el censo a un archivo .xlsx con openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Módulo no instalado",
                "Para exportar a Excel necesitas instalar openpyxl.\n\n"
                "Abre una terminal y ejecuta:\n    pip install openpyxl\n\nLuego vuelve a intentarlo.")
            return

        datos = cargar_datos(CONFIG['archivo_datos'])
        if not datos:
            messagebox.showwarning("Aviso", "El censo está vacío.")
            return

        import unicodedata
        datos.sort(key=lambda x: unicodedata.normalize('NFD', x.get('nombre', '')).encode('ascii','ignore').decode().lower())

        archivo = filedialog.asksaveasfilename(
            title="Guardar Excel del Censo",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Censo_Agonia.xlsx"
        )
        if not archivo:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Censo Costaleros"

        color_morado = "4F1243"
        color_oro    = "D4AF37"
        fill_cabecera = PatternFill("solid", fgColor=color_oro)
        fill_alterno  = PatternFill("solid", fgColor="F9F9F9")
        font_cabecera = Font(bold=True, color="333333", size=11)
        font_titulo   = Font(bold=True, color=color_morado, size=13)

        ws.merge_cells("A1:G1")
        ws["A1"] = "Censo de Costaleros — OFS Mayordomía de Ntro. Padre Jesús Nazareno"
        ws["A1"].font = font_titulo
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y')}   |   Total: {len(datos)} costaleros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(italic=True, color="888888", size=10)

        cabeceras = ["ID", "Nombre", "Altura (cm)", "Teléfono", "Hombro preferido", "Miércoles Santo", "Viernes Santo"]
        for col, cab in enumerate(cabeceras, 1):
            c = ws.cell(row=4, column=col, value=cab)
            c.font = font_cabecera
            c.fill = fill_cabecera
            c.alignment = Alignment(horizontal="center")

        for fila, p in enumerate(datos, 5):
            valores = [
                p.get('id'),
                p.get('nombre', ''),
                p.get('altura', 0),
                p.get('telefono', '') or '',
                (p.get('pref_hombro', '') or 'Indiferente').capitalize(),
                "Sí" if p.get('miercoles_santo') else "No",
                "Sí" if p.get('viernes_santo') else "No",
            ]
            for col, val in enumerate(valores, 1):
                c = ws.cell(row=fila, column=col, value=val)
                c.alignment = Alignment(horizontal="center" if col != 2 else "left")
                if fila % 2 == 0:
                    c.fill = fill_alterno

        anchos = [6, 35, 12, 16, 18, 16, 14]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

        try:
            wb.save(archivo)
            self._ultimo_formato = "Excel"
            self.mostrar_toast("✅  Excel guardado correctamente", color="#27ae60")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")

    # --- PANTALLA CENSO ---
    def crear_pantalla_censo(self):
        self._censo_sort_col = "Nombre"
        self._censo_sort_rev = False
        self._ultimo_formato = "PDF"   

        f = tk.Frame(self.frame_main, bg=C_GRIS_FONDO, padx=30, pady=30)
        tk.Label(f, text="Gestión del Censo General", font=("Segoe UI", 22, "bold"), bg=C_GRIS_FONDO, fg=C_MORADO).pack(anchor="w")
        
        toolbar = tk.Frame(f, bg=C_GRIS_FONDO)
        toolbar.pack(fill=tk.X, pady=15)
        
        btn_nuevo = self.crear_boton_moderno(toolbar, "➕ Nuevo", "#27ae60", "#2ecc71", C_BLANCO, command=lambda: self.abrir_formulario_costalero())
        btn_nuevo.pack(side=tk.LEFT, padx=(0, 10))
        btn_editar = self.crear_boton_moderno(toolbar, "✏️ Editar", C_ORO, C_ORO_HOVER, C_TEXTO, command=lambda: self.abrir_formulario_costalero(editar=True))
        btn_editar.pack(side=tk.LEFT, padx=10)
        btn_borrar = self.crear_boton_moderno(toolbar, "❌ Borrar", "#ff4757", "#ff6b81", C_BLANCO, command=self.borrar_costalero)
        btn_borrar.pack(side=tk.LEFT, padx=10)
        
        # --- BOTONES DE LA NUBE ---
        btn_subir = self.crear_boton_moderno(toolbar, "☁️ Subir a Nube", "#2980b9", "#3498db", C_BLANCO, command=self.subir_a_nube)
        btn_subir.pack(side=tk.LEFT, padx=10)
        
        btn_bajar = self.crear_boton_moderno(toolbar, "📥 Bajar de Nube", "#8e44ad", "#9b59b6", C_BLANCO, command=self.descargar_de_nube)
        btn_bajar.pack(side=tk.LEFT, padx=10)
        
        search_frame = tk.Frame(toolbar, bg=C_GRIS_FONDO)
        search_frame.pack(side=tk.RIGHT)
        tk.Label(search_frame, text="🔍 Buscar:", bg=C_GRIS_FONDO, font=("Segoe UI", 11)).pack(side=tk.LEFT, padx=(0,5))
        self.entry_busqueda = tk.Entry(search_frame, font=("Segoe UI", 11), width=25, relief="solid", bd=1)
        self.entry_busqueda.pack(side=tk.LEFT)
        self.entry_busqueda.bind("<KeyRelease>", self.actualizar_tabla_censo)

        def exportar_ultimo():
            if self._ultimo_formato == "Excel":
                self.exportar_censo_excel()
            else:
                self.exportar_censo_pdf()

        def mostrar_menu_exportar(event=None):
            menu_exp = tk.Menu(self.root, tearoff=0)
            menu_exp.add_command(label="📄  Exportar como PDF",   command=self.exportar_censo_pdf)
            menu_exp.add_command(label="📊  Exportar como Excel", command=self.exportar_censo_excel)
            w = btn_exp_main
            menu_exp.post(w.winfo_rootx(), w.winfo_rooty() + w.winfo_height())

        split_frame = tk.Frame(toolbar, bg=C_MORADO, cursor="hand2")
        split_frame.pack(side=tk.LEFT, padx=10)

        btn_exp_main = tk.Button(split_frame, text="📤  Exportar", bg=C_MORADO, fg=C_BLANCO,
                                  font=("Segoe UI", 11, "bold"), bd=0, padx=14, pady=10,
                                  activebackground=C_MORADO_HOVER, activeforeground=C_BLANCO,
                                  cursor="hand2", command=exportar_ultimo)
        btn_exp_main.pack(side=tk.LEFT)

        tk.Frame(split_frame, width=1, bg=C_MORADO_HOVER).pack(side=tk.LEFT, fill=tk.Y, pady=4)

        btn_exp_arrow = tk.Button(split_frame, text="▼", bg=C_MORADO, fg=C_BLANCO,
                                   font=("Segoe UI", 9), bd=0, padx=8, pady=10,
                                   activebackground=C_MORADO_HOVER, activeforeground=C_BLANCO,
                                   cursor="hand2", command=mostrar_menu_exportar)
        btn_exp_arrow.pack(side=tk.LEFT)

        filtros_frame = tk.Frame(f, bg=C_GRIS_FONDO)
        filtros_frame.pack(fill=tk.X, pady=(0, 8))

        tk.Label(filtros_frame, text="Filtrar:", bg=C_GRIS_FONDO, font=("Segoe UI", 10, "bold"), fg="#555").pack(side=tk.LEFT, padx=(0, 8))
        self._var_filtro_mi = tk.BooleanVar(value=False)
        self._var_filtro_vi = tk.BooleanVar(value=False)
        chk_mi = ttk.Checkbutton(filtros_frame, text="Solo Miércoles Santo", variable=self._var_filtro_mi, command=self.actualizar_tabla_censo)
        chk_mi.pack(side=tk.LEFT, padx=(0, 15))
        chk_vi = ttk.Checkbutton(filtros_frame, text="Solo Viernes Santo", variable=self._var_filtro_vi, command=self.actualizar_tabla_censo)
        chk_vi.pack(side=tk.LEFT)

        self._lbl_contador = tk.Label(filtros_frame, text="", bg=C_GRIS_FONDO, font=("Segoe UI", 10), fg="#555")
        self._lbl_contador.pack(side=tk.RIGHT)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", font=('Segoe UI', 11, 'bold'), background=C_ORO, foreground=C_TEXTO, relief="flat")
        style.configure("Treeview", font=('Segoe UI', 11), rowheight=30, background=C_BLANCO, fieldbackground=C_BLANCO, borderwidth=0)
        style.map('Treeview', background=[('selected', C_MORADO)], foreground=[('selected', C_BLANCO)])

        frame_tabla = tk.Frame(f, bg=C_BLANCO, highlightbackground="#e0e0e0", highlightthickness=1)
        frame_tabla.pack(fill=tk.BOTH, expand=True)

        columnas = ("ID", "Nombre", "Telefono", "Altura", "Hombro", "Miércoles", "Viernes")
        self.tree = ttk.Treeview(frame_tabla, columns=columnas, show="headings")

        def cmd_sort(col):
            if self._censo_sort_col == col:
                self._censo_sort_rev = not self._censo_sort_rev
            else:
                self._censo_sort_col = col
                self._censo_sort_rev = False
            self.actualizar_tabla_censo()

        self.tree.heading("ID",        text="ID",                  command=lambda: cmd_sort("ID"))
        self.tree.heading("Nombre",    text="Nombre del Costalero",command=lambda: cmd_sort("Nombre"))
        self.tree.heading("Telefono",  text="Teléfono")
        self.tree.heading("Altura",    text="Altura (cm)",         command=lambda: cmd_sort("Altura"))
        self.tree.heading("Hombro",    text="Preferencia")
        self.tree.heading("Miércoles", text="M. Santo",            command=lambda: cmd_sort("Miércoles"))
        self.tree.heading("Viernes",   text="V. Santo",            command=lambda: cmd_sort("Viernes"))

        self.tree.column("ID",        width=50,  anchor="center")
        self.tree.column("Nombre",    width=220)
        self.tree.column("Telefono",  width=120, anchor="center")
        self.tree.column("Altura",    width=100, anchor="center")
        self.tree.column("Hombro",    width=120, anchor="center")
        self.tree.column("Miércoles", width=100, anchor="center")
        self.tree.column("Viernes",   width=100, anchor="center")
        
        scrollbar = ttk.Scrollbar(frame_tabla, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        return f

    def actualizar_tabla_censo(self, event=None):
        import unicodedata
        def normalizar(s):
            return unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().lower()

        for item in self.tree.get_children():
            self.tree.delete(item)
            
        datos = cargar_datos(CONFIG['archivo_datos'])
        filtro_texto = normalizar(self.entry_busqueda.get()) if hasattr(self, 'entry_busqueda') else ""

        filtro_mi = self._var_filtro_mi.get() if hasattr(self, '_var_filtro_mi') else False
        filtro_vi = self._var_filtro_vi.get() if hasattr(self, '_var_filtro_vi') else False
        if filtro_mi: datos = [p for p in datos if p.get('miercoles_santo')]
        if filtro_vi: datos = [p for p in datos if p.get('viernes_santo')]

        if filtro_texto:
            datos = [p for p in datos if
                     filtro_texto in normalizar(p.get('nombre', ''))
                     or str(p.get('id', '')) == filtro_texto
                     or str(p.get('altura', '')) == filtro_texto]

        col  = getattr(self, '_censo_sort_col', 'Nombre')
        rev  = getattr(self, '_censo_sort_rev', False)
        if col == "ID":          datos.sort(key=lambda x: x.get('id', 0),                    reverse=rev)
        elif col == "Nombre":    datos.sort(key=lambda x: normalizar(x.get('nombre', '')),    reverse=rev)
        elif col == "Altura":    datos.sort(key=lambda x: x.get('altura', 0),                 reverse=rev)
        elif col == "Miércoles": datos.sort(key=lambda x: x.get('miercoles_santo', False),    reverse=rev)
        elif col == "Viernes":   datos.sort(key=lambda x: x.get('viernes_santo', False),      reverse=rev)

        etiquetas = {
            "ID":        "ID",
            "Nombre":    "Nombre del Costalero",
            "Altura":    "Altura (cm)",
            "Miércoles": "M. Santo",
            "Viernes":   "V. Santo",
        }
        for c, texto_base in etiquetas.items():
            if c == col:
                indicador = " ▼" if rev else " ▲"
                self.tree.heading(c, text=texto_base + indicador)
            else:
                self.tree.heading(c, text=texto_base)

        for p in datos:
            mi = "✅" if p.get("miercoles_santo") else "❌"
            vi = "✅" if p.get("viernes_santo") else "❌"
            hombro = p.get("pref_hombro", "") or "Indiferente"
            telefono = p.get("telefono", "")
            self.tree.insert("", tk.END, values=(p['id'], p['nombre'], telefono, p['altura'], hombro.capitalize(), mi, vi))

        if hasattr(self, '_lbl_contador'):
            total_censo = cargar_datos(CONFIG['archivo_datos'])
            total     = len(total_censo)
            n_mi      = sum(1 for p in total_censo if p.get('miercoles_santo'))
            n_vi      = sum(1 for p in total_censo if p.get('viernes_santo'))
            mostrando = len(datos)
            if mostrando < total:
                texto = f"Mostrando {mostrando} de {total}  |  🕯️ Miércoles: {n_mi}  |  ✝️ Viernes: {n_vi}"
            else:
                texto = f"Total: {total} costaleros  |  🕯️ Miércoles: {n_mi}  |  ✝️ Viernes: {n_vi}"
            self._lbl_contador.config(text=texto)

    def abrir_formulario_costalero(self, editar=False):
        import unicodedata
        def norm(s): return unicodedata.normalize('NFD', str(s)).encode('ascii','ignore').decode().lower().strip()

        datos = cargar_datos(CONFIG['archivo_datos'])
        costalero = None
        
        if editar:
            seleccion = self.tree.selection()
            if not seleccion:
                messagebox.showwarning("Atención", "Selecciona un costalero de la lista para editar.")
                return
            pid = self.tree.item(seleccion[0])['values'][0]
            costalero = next((x for x in datos if x['id'] == pid), None)
            if not costalero: return

        top = tk.Toplevel(self.root)
        top.title("Editar Costalero" if editar else "Nuevo Costalero")
        top.geometry("400x530")
        top.configure(bg=C_BLANCO)
        top.resizable(True, True)
        top.transient(self.root) 
        top.grab_set()

        top.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (400 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (530 // 2)
        top.geometry(f"+{x}+{y}")

        tk.Label(top, text="📋 Ficha del Costalero", font=("Segoe UI", 16, "bold"), bg=C_BLANCO, fg=C_MORADO).pack(pady=(20, 20))

        form_frame = tk.Frame(top, bg=C_BLANCO)
        form_frame.pack(padx=40, fill=tk.BOTH, expand=True)

        tk.Label(form_frame, text="Nombre Completo:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        var_nombre = tk.StringVar(value=costalero['nombre'] if costalero else "")
        tk.Entry(form_frame, textvariable=var_nombre, font=("Segoe UI", 12), relief="solid", bd=1).pack(fill=tk.X, pady=(2, 10))

        tk.Label(form_frame, text="Teléfono:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        var_telefono = tk.StringVar(value=costalero.get('telefono', '') if costalero else "")
        tk.Entry(form_frame, textvariable=var_telefono, font=("Segoe UI", 12), relief="solid", bd=1).pack(fill=tk.X, pady=(2, 10))

        tk.Label(form_frame, text="Altura de hombro (en cm):", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        var_altura = tk.StringVar(value=str(costalero['altura']) if costalero else "")
        tk.Entry(form_frame, textvariable=var_altura, font=("Segoe UI", 12), relief="solid", bd=1).pack(fill=tk.X, pady=(2, 10))

        tk.Label(form_frame, text="Preferencia de Hombro:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        var_hombro = tk.StringVar(value=costalero.get('pref_hombro', 'Indiferente') if costalero and costalero.get('pref_hombro') else "Indiferente")
        combo_hombro = ttk.Combobox(form_frame, textvariable=var_hombro, values=["Derecho", "Izquierdo", "Ambos", "Indiferente"], state="readonly", font=("Segoe UI", 11))
        combo_hombro.pack(fill=tk.X, pady=(2, 15))

        tk.Label(form_frame, text="Disponibilidad para procesionar:", bg=C_BLANCO, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        var_miercoles = tk.BooleanVar(value=costalero.get('miercoles_santo', True) if costalero else True)
        var_viernes   = tk.BooleanVar(value=costalero.get('viernes_santo',   True) if costalero else True)
        chk_style = ttk.Style()
        chk_style.configure("TCheckbutton", background=C_BLANCO, font=("Segoe UI", 11))
        ttk.Checkbutton(form_frame, text="Sale el Miércoles Santo", variable=var_miercoles).pack(anchor="w", pady=2)
        ttk.Checkbutton(form_frame, text="Sale el Viernes Santo",   variable=var_viernes).pack(anchor="w", pady=2)

        def guardar():
            nombre = var_nombre.get().strip()
            altura_str = var_altura.get().strip()

            if not nombre or not altura_str.isdigit():
                messagebox.showwarning("Error", "Revisa los campos. El nombre no puede estar vacío y la altura debe ser un número.")
                return

            tel = var_telefono.get().strip()
            if tel:
                tel_limpio = tel.replace(" ", "").replace("-", "").replace("+", "")
                if not tel_limpio.isdigit() or len(tel_limpio) < 9:
                    messagebox.showwarning("Teléfono incorrecto",
                        "El teléfono no parece válido.\nDebe tener al menos 9 dígitos y solo puede contener números, espacios, guiones o el signo +.")
                    return

            altura_val = int(altura_str)
            if altura_val < 100 or altura_val > 200:
                messagebox.showerror("Altura inválida",
                    f"La altura '{altura_val} cm' no parece correcta.\nRecuerda que es la altura de hombro, que suele estar entre 120 y 175 cm aproximadamente.")
                return
            if not (115 <= altura_val <= 180):
                continuar = messagebox.askyesno("Altura atípica",
                    f"La altura de hombro '{altura_val} cm' está fuera del rango habitual (115–180 cm).\n\n¿Es correcta y deseas guardarla igualmente?")
                if not continuar:
                    return

            if not editar:
                nombre_norm = norm(nombre)
                for p in datos:
                    if norm(p.get('nombre', '')) == nombre_norm:
                        continuar = messagebox.askyesno("⚠️ Posible duplicado",
                            f"Ya existe un costalero con un nombre muy similar:\n\n  → {p['nombre']} (ID {p['id']})\n\n¿Estás seguro de que es una persona distinta y quieres añadirlo?")
                        if not continuar:
                            return
                        break

            hombro_val = var_hombro.get()
            if hombro_val == "Indiferente": hombro_val = ""

            if editar:
                costalero['nombre']          = nombre
                costalero['telefono']        = tel
                costalero['altura']          = altura_val
                costalero['pref_hombro']     = hombro_val
                costalero['miercoles_santo'] = var_miercoles.get()
                costalero['viernes_santo']   = var_viernes.get()
            else:
                nuevo_id = max([p.get('id', 0) for p in datos], default=0) + 1
                datos.append({
                    "id": nuevo_id, "nombre": nombre, "telefono": tel,
                    "altura": altura_val, "pref_hombro": hombro_val,
                    "puede_repetir": True,
                    "miercoles_santo": var_miercoles.get(),
                    "viernes_santo":   var_viernes.get()
                })

            if self.guardar_censo(datos):
                self.actualizar_tabla_censo()
                top.destroy()
                self.mostrar_ficha_guardada(
                    nombre, altura_val, tel,
                    hombro_val, var_miercoles.get(), var_viernes.get(),
                    editar=editar
                )

        top.bind("<Return>",  lambda e: guardar())
        top.bind("<Escape>",  lambda e: top.destroy())

        btn_guardar = tk.Button(top, text="💾 GUARDAR CAMBIOS", bg=C_MORADO, fg=C_BLANCO,
                                 font=("Segoe UI", 12, "bold"), bd=0, cursor="hand2", command=guardar)
        btn_guardar.pack(fill=tk.X, padx=40, pady=(20, 0), ipady=8)

    def borrar_costalero(self):
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Selecciona un costalero de la lista haciendo clic sobre él.")
            return
            
        item = self.tree.item(seleccion[0])
        pid = item['values'][0]
        nombre = item['values'][1]
        
        if messagebox.askyesno("⚠️ Confirmar Borrado", f"¿Estás completamente seguro de querer dar de baja a {nombre}?\n\nEsta acción no se puede deshacer."):
            datos = cargar_datos(CONFIG['archivo_datos'])
            datos = [x for x in datos if x.get('id') != pid]
            self.guardar_censo(datos)
            self.actualizar_tabla_censo()

    # ==========================================
    # SINCRONIZACIÓN EN LA NUBE (FIREBASE)
    # ==========================================
    def subir_a_nube(self):
        if not messagebox.askyesno("⚠️ Confirmar Subida", "Vas a subir tu censo actual a internet.\nA partir de ahora, esta será la versión oficial para todos los ordenadores.\n\n¿Deseas continuar?"):
            return

        try:
            with open(CONFIG['archivo_datos'], 'r', encoding='utf-8') as f:
                datos_locales = json.load(f)
            
            respuesta = requests.put(FIREBASE_URL, json=datos_locales, timeout=10)
            
            if respuesta.status_code == 200:
                messagebox.showinfo("Sincronización", "☁️ ¡Datos subidos correctamente a la nube!\n\nTu censo local es ahora la versión oficial en internet.")
            else:
                messagebox.showerror("Error", f"Hubo un fallo al subir a la nube. Código: {respuesta.status_code}")
                
        except Exception as e:
            messagebox.showerror("Error de Conexión", f"No se pudo conectar a internet o a Firebase.\nDetalles: {e}")

    def descargar_de_nube(self):
        if not messagebox.askyesno("⚠️ Confirmar Descarga", "Vas a descargar el censo oficial de internet.\nEsto SOBREESCRIBIRÁ el censo que tienes actualmente en este ordenador.\n\n(Se creará una copia de seguridad automática por si acaso).\n\n¿Estás seguro de continuar?"):
            return
            
        try:
            respuesta = requests.get(FIREBASE_URL, timeout=10)
            
            if respuesta.status_code == 200:
                datos_nube = respuesta.json()
                
                if datos_nube:
                    archivo_actual = CONFIG['archivo_datos']
                    
                    # 1. Crear copia de seguridad local antes de machacar
                    if os.path.exists(archivo_actual):
                        fecha_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        shutil.copy(archivo_actual, f"censo_backup_pre_nube_{fecha_str}.json")
                        
                    # 2. Guardar los datos de la nube
                    with open(archivo_actual, 'w', encoding='utf-8') as f:
                        json.dump(datos_nube, f, indent=4, ensure_ascii=False)
                        
                    # 3. Refrescar la tabla para ver los cambios
                    self.actualizar_tabla_censo() 
                    messagebox.showinfo("Sincronización", "📥 ¡Censo descargado de la nube con éxito y actualizado en pantalla!")
                else:
                    messagebox.showwarning("Aviso", "La base de datos en la nube está vacía.\nSube los datos primero desde el ordenador que tenga el censo oficial.")
            else:
                messagebox.showerror("Error", f"Fallo al descargar de la nube. Código: {respuesta.status_code}")
                
        except Exception as e:
            messagebox.showerror("Error de Conexión", f"No se pudo conectar a internet o a Firebase.\nDetalles: {e}")

# ==========================================
# INICIO DE LA APLICACIÓN
# ==========================================
if __name__ == "__main__":
    root = tk.Tk()
    app = GestorCofradeAPP(root)
    
    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    
    root.mainloop()